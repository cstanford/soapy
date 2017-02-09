'''
*************************************************
* Soapy                                         *
*                                               *
* A Simple script for creating SLP SOAP notes   *
* in ms excel                                   *
*************************************************
'''

import os
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

parser = argparse.ArgumentParser()
parser.add_argument('filename')
args = parser.parse_args()
ft_heading = Font(size=11, bold=True)

# column headings
header_cells = {'A': 'Date', 'B': 'Subjective Notes', 'C': 'Short-term Goals', 'D': 'Data (+/-)', 'E': 'Cues Given', 'F': 'Assessment/Plan'}

# set appropriate column widths
column_widths = {'A': 10, 'B': 20, 'C': 35, 'D': 24, 'E': 20, 'F': 20}

# set up print options. ws = worksheet
def print_ops_setup(ws):	
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
	ws.print_options.gridLines = True
	ws.page_margins.left = 0.25
	ws.page_margins.right = 0.25
	ws.page_margins.bottom = 0.5

# make cell heading
def insert_column_headings(ws, startRow):
	for k, v in header_cells.items():
		cell = k + str(startRow)
		ws[cell].font = ft_heading
		ws[cell].value = v

# insert goals to cells
def insert_patient_goals(ws, startRow, goals):
	for goal in range(startRow, len(goals) + startRow):
		cell_num = 'C' + str(goal)
		cell = ws[cell_num].value = goals[goal - startRow]

def input_data(ws, goal_index, heading_index, offset, goals):
	for i in range(0, 4):
		ws.merge_cells('B' + str(goal_index) +':B' + str(goal_index + offset - 1)) # merges subjective note cells
		insert_column_headings(ws, heading_index)
		insert_patient_goals(ws, goal_index, goals)
		heading_index += offset + 3
		goal_index += offset + 3


def soapy(wb, is_new_wb, dest_filename):

	patient_name = input('Patient name: ' )

	if is_new_wb:
		ws = wb.active
		ws.title = patient_name
	else:
		if patient_name in wb.sheetnames: # delete patient's worksheet if it already exists
			while(True):
				print('Patient: \'' + patient_name + '\' already exists.')
				selection = input('Enter (1) to keep existing patient or (2) to overwrite: ')
				selection = selection.replace(' ', '')
				print('\n')
				
				if selection == '2':
					wb.remove(wb[patient_name])
					break
				elif selection == '1':
					break
				else:
					print('Invalid input. Please enter (1) or (2).')

		ws = wb.create_sheet(title=patient_name) # create a new sheet for the patient. 


	print_ops_setup(ws)

	while True:
		try:
			num_goals = int(input('Number of goals: ' ))
			break;
		except ValueError:
			print('Invalid input. Please enter an integer.')

	goals = []

	# get goals from user
	for i in range(0,num_goals):
		goal = input('Goal ' + str(i + 1) + ': ')
		goals.append(goal)

	# merge top row of cells and input patient name
	ws.merge_cells('A1:F1')
	a1 = ws['A1']
	a1.font = ft_heading
	a1.value = 'Pt: ' + patient_name

	for column, width in column_widths.items():
		ws.column_dimensions[column].width = width

	heading_index = 4 # where to start inserting heading row
	goal_index = 5	# where to start inserting goals

	custom_offset = 4 # Use if num_goals < 4 to make room for subjective notes. 
	use_custom_offset = False

	# If num_goals < 4, we use a custom offset.
	# The custom offset guarantees that the spreadsheet will have enough room
	# to take subjective notes. 
	# In other words, we will have at least four merged rows in column B to 
	# take notes. 
	if num_goals < custom_offset:
		use_custom_offset = True

	if use_custom_offset == True:
		input_data(ws, goal_index,heading_index,custom_offset, goals)
	else:
		input_data(ws, goal_index,heading_index,num_goals, goals)


	while True:
		try:
			wb.save(filename=dest_filename)
			print('\nSuccessfully updated ' + dest_filename + '.\n')
			break
		except:
			print('Failed to save file.')
			print('This is probalbly because the file is currently open in excel.')
			print('Please close the file and try again.\n')
			selection = input('Press any key to try again. Press (0) to quit application: ')
			if selection == '0':
				exit()


def main():
	filename = args.filename
	dest_dir = 'workbooks'
	dest_filename = dest_dir + '/' + filename + '.xlsx'

	# create a workbooks directory if one does not exist
	if os.path.isdir(dest_dir) == False:
		os.makedirs(dest_dir)

	if os.path.isfile(dest_filename): # open workbook if exists
		wb = load_workbook(filename = dest_filename)
		is_new_wb = False
	else:
		wb = Workbook()
		is_new_wb = True

	print('\n***********  SOAPY  ***********')
	print('Be sure that excel is not open!\n')

	while True:
		soapy(wb, is_new_wb, dest_filename)
		# soapy() has already been called so we know that we are
		# not dealing with a new wb
		is_new_wb = False
		print('Would you add another patient to this workbook?')
		selection = input('(y/n): ')
		print('\n')

		if selection == 'y':
			continue
		else:
			print('Goodbye <3.')
			print('*******************************\n')
			break

main()

